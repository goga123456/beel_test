import asyncio
import time
from datetime import datetime
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from sender import *
from aiogram import types, executor, Bot, Dispatcher
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
import schedule
from openpyxl import load_workbook
import states
from markups.keyboard import *
from markups.markup_kalendar import get_birthday_kb, get_birthday_month_kb, get_birthday_day_kb, get_birthday_year_kb
from markups.reply_markups_start_and_back import get_start_kb, get_start_and_back_kb
from messages import *
from states import ProfileStatesGroup

from config import TOKEN_API

storage = MemoryStorage()
bot = Bot(TOKEN_API)
dp = Dispatcher(bot,
                storage=storage)
scheduler = AsyncIOScheduler()

scheduler.add_job(send_email, "cron", hour=17, minute=40)
@dp.message_handler(commands=['start'], state='*')
async def cmd_start(message: types.Message, state: FSMContext) -> None:
    await bot.send_message(chat_id=message.from_user.id,
                           text=start_msg,
                           reply_markup=get_initial_kb())
    if state is None:
        return
    await state.finish()


@dp.message_handler(content_types=['text'], state=ProfileStatesGroup.cause_of_rejection)
async def load_it_info(message: types.Message, state: FSMContext) -> None:
    async with state.proxy() as data:
        data['cause'] = message.text

        now = datetime.now()
        response_date = now.strftime("%d.%m.%Y %H:%M:%S")

        wb = load_workbook("example.xlsx")
        ws = wb['Лист1']
        ws.append([response_date, data['cause']])
        wb.save("example.xlsx")
        wb.close()

    await bot.send_message(chat_id=message.from_user.id,
                           text=again)
    await state.finish()


@dp.message_handler(content_types=['text'], state=ProfileStatesGroup.input_number)
async def load_it_info(message: types.Message, state: FSMContext) -> None:
    async with state.proxy() as data:
        data['number'] = message.text
        if str(data['number']).isdigit() and str(data['number']).startswith('99893'):
            if len(str(data['number'])) == 12:
                await bot.send_message(chat_id=message.from_user.id,
                                       text=name, reply_markup=get_start_and_back_kb())
                await ProfileStatesGroup.input_name.set()
        else:
            await bot.send_message(chat_id=message.from_user.id,
                                   text=wrong_number)


@dp.message_handler(content_types=['text'], state=ProfileStatesGroup.input_name)
async def load_it_info(message: types.Message, state: FSMContext) -> None:
    async with state.proxy() as data:
        data['name'] = message.text
    if message.text == 'Назад':
        await bot.send_message(chat_id=message.from_user.id,
                               text=number,
                               reply_markup=get_start_kb())
        await ProfileStatesGroup.input_number.set()
    else:
        await bot.send_message(chat_id=message.from_user.id,
                               text=surname, reply_markup=get_start_and_back_kb())
        await ProfileStatesGroup.input_surname.set()






@dp.message_handler(content_types=['text'], state=ProfileStatesGroup.input_surname)
async def load_it_info(message: types.Message, state: FSMContext) -> None:
    async with state.proxy() as data:
        data['surname'] = message.text
        data['day'] = '-'
        data['month'] = '-'
        data['year'] = '-'
    if message.text == 'Назад':
        await bot.send_message(chat_id=message.from_user.id,
                               text=name)
        await ProfileStatesGroup.input_name.set()
    else:
        await bot.send_message(chat_id=message.from_user.id,
                               text=date_of_birthday,
                               reply_markup=get_birthday_kb())
        await ProfileStatesGroup.input_birthday.set()


@dp.message_handler(content_types=['text'], state=ProfileStatesGroup.input_other_town_and_district)
async def load_it_info(message: types.Message, state: FSMContext) -> None:
    async with state.proxy() as data:
        data['town_and_district'] = message.text
    if message.text == 'Назад':
        await bot.send_message(chat_id=message.from_user.id,
                               text=where_are_you_from,
                               reply_markup=get_town_kb())
        await ProfileStatesGroup.input_Tashkent_or_other_town.set()
    else:
        await bot.send_message(chat_id=message.from_user.id,
                               text=education,
                               reply_markup=get_edu_kb())
        await ProfileStatesGroup.input_edu.set()

@dp.message_handler(content_types=['text'], state=ProfileStatesGroup.experience_describe)
async def load_it_info(message: types.Message, state: FSMContext) -> None:
    async with state.proxy() as data:
        data['exp'] = message.text
    if message.text == 'Назад':
        await bot.send_message(chat_id=message.from_user.id,
                               text=experience_msg,
                               reply_markup=get_exp_kb())
        await ProfileStatesGroup.input_experience.set()
    else:
        await bot.send_message(chat_id=message.from_user.id,
                               text=time_for_call)
        await ProfileStatesGroup.input_day_and_time.set()


@dp.message_handler(content_types=['text'], state=ProfileStatesGroup.input_day_and_time)
async def load_it_info(message: types.Message, state: FSMContext) -> None:
    async with state.proxy() as data:
        data['day_and_time'] = message.text
    if message.text == 'Назад':
        await bot.send_message(chat_id=message.from_user.id,
                               text=experience_msg,
                               reply_markup=get_exp_kb())
        await ProfileStatesGroup.input_experience.set()
    else:
        await bot.send_message(chat_id=message.from_user.id,
                               text=thank_you)
        await bot.send_message(chat_id=message.from_user.id,
                               text=sendmail)
        await bot.send_message(chat_id=message.from_user.id,
                               text=again,
                               reply_markup=get_start_kb())
        now = datetime.now()
        response_date = now.strftime("%d.%m.%Y %H:%M:%S")
        without_spaces = str(data['month']).replace(" ", "")
        birthday = f"{data['day']}.{without_spaces}.{data['year']}"
        wb = load_workbook("example.xlsx")
        ws = wb['Лист1']
        ws.append([response_date, '', data['number'], data['name'], data['surname'], birthday, data['town_and_district'], data['edu'], data['rus'], data['uzb'], data['eng'], data['exp'], data['day_and_time']])
        wb.save("example.xlsx")
        wb.close()
        await state.finish()



@dp.callback_query_handler()
async def initial_keyboards(callback_query: types.CallbackQuery):
    if callback_query.data == 'next':
        await callback_query.message.delete()
        await bot.send_message(chat_id=callback_query.from_user.id,
                               text=start_msg2,
                               reply_markup=get_initial_kb2())
    if callback_query.data == 'close':
        await ProfileStatesGroup.cause_of_rejection.set()
        await bot.send_message(callback_query.from_user.id, text=cause_of_rejection)
        await callback_query.message.delete()
    if callback_query.data == 'yes_i_want':
        await ProfileStatesGroup.input_number.set()
        await bot.send_message(callback_query.from_user.id, text=number, reply_markup=get_start_kb())
        await callback_query.message.delete()
    if callback_query.data == 'i_dont_want':
        await ProfileStatesGroup.cause_of_rejection.set()
        await bot.send_message(callback_query.from_user.id, text=cause_of_rejection)
        await callback_query.message.delete()


# колбеки на первые 2 сообщения


@dp.callback_query_handler(state=ProfileStatesGroup.input_birthday)
async def calendar_keyboard(callback_query: types.CallbackQuery, state: FSMContext):
    if callback_query.data == 'day':
        await bot.edit_message_text(chat_id=callback_query.message.chat.id,
                                    message_id=callback_query.message.message_id,
                                    text=choose_day, reply_markup=get_birthday_day_kb())
        await ProfileStatesGroup.input_day.set()
    if callback_query.data == 'month':
        await bot.edit_message_text(chat_id=callback_query.message.chat.id,
                                    message_id=callback_query.message.message_id,
                                    text=choose_month, reply_markup=get_birthday_month_kb())
        await ProfileStatesGroup.input_month.set()
    if callback_query.data == 'year':
        await bot.edit_message_text(chat_id=callback_query.message.chat.id,
                                    message_id=callback_query.message.message_id,
                                    text=choose_year, reply_markup=get_birthday_year_kb())
        await ProfileStatesGroup.input_year.set()
    if callback_query.data == 'send_birth':
        async with state.proxy() as data:
            if data['day'] == '-' and data['month'] == '-' and data['year'] == '-':
                await bot.send_message(callback_query.message.chat.id, "Дата не выбрана")
            elif data['day'] == '-' and data['month'] == '-':
                await bot.send_message(callback_query.message.chat.id, "День и месяц не выбраны")
            elif data['day'] == '-' and data['year'] == '-':
                await bot.send_message(callback_query.message.chat.id, "День и год не выбраны")
            elif data['month'] == '-' and data['year'] == '-':
                await bot.send_message(callback_query.message.chat.id, "Месяц и год не выбраны")
            elif data['month'] == '-' and data['year'] == '-':
                await bot.send_message(callback_query.message.chat.id, "Месяц и год не выбраны")
            elif data['day'] == '-':
                await bot.send_message(callback_query.message.chat.id, "День не выбран")
            elif data['month'] == '-':
                await bot.send_message(callback_query.message.chat.id, "Месяц не выбран")
            elif data['year'] == '-':
                await bot.send_message(callback_query.message.chat.id, text="Год не выбран")
            elif data['month'] == '0 2' and data['day'] == '30':
                await bot.send_message(callback_query.message.chat.id, text=data_not_exist)
            elif data['month'] == '0 2' and data['day'] == '31':
                await bot.send_message(callback_query.message.chat.id, text=data_not_exist)
            elif data['month'] == '0 4' and data['day'] == '31':
                await bot.send_message(callback_query.message.chat.id, text=data_not_exist)
            elif data['month'] == '0 6' and data['day'] == '31':
                await bot.send_message(callback_query.message.chat.id, text=data_not_exist)
            elif data['month'] == '0 9' and data['day'] == '31':
                await bot.send_message(callback_query.message.chat.id, text=data_not_exist)
            elif data['month'] == '1 1' and data['day'] == '31':
                await bot.send_message(callback_query.message.chat.id, text=data_not_exist)
            else:
                without_spaces = str(data['month']).replace(" ", "")
                now = datetime.now()
                response_date = now.strftime("%d.%m.%Y %H:%M:%S")
                birthday = f"{data['day']}.{without_spaces}.{data['year']}"
                if now.year - int(data['year']) < 18:
                    await bot.send_message(callback_query.message.chat.id, text=less_than_18)
                    await bot.send_message(callback_query.message.chat.id, text=again)

                    wb = load_workbook("example.xlsx")
                    ws = wb['Лист1']
                    ws.append([response_date, '', data['number'], data['name'], data['surname'], birthday])
                    wb.save("example.xlsx")
                    wb.close()
                    ###Добавление в базу данных

                    await callback_query.message.delete()
                    await state.finish()
                else:
                    await callback_query.message.delete()
                    await bot.send_message(chat_id=callback_query.message.chat.id,
                                           text=date_of_birthday)
                    await bot.send_message(callback_query.from_user.id,
                                           text=f"{data['day']}.{without_spaces}.{data['year']}")
                    await bot.send_message(chat_id=callback_query.message.chat.id,
                                           text=where_are_you_from,
                                           reply_markup=get_town_kb())
                    await states.ProfileStatesGroup.input_Tashkent_or_other_town.set()

    if callback_query.data == 'back_to_surname':
        await callback_query.message.delete()
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=surname,
                               reply_markup=get_start_and_back_kb())
        await ProfileStatesGroup.input_surname.set()


@dp.callback_query_handler(state=ProfileStatesGroup.input_day)
async def day_keyboard(callback_query: types.CallbackQuery, state: FSMContext):
    if (callback_query.data == '1' or '2' or '3' or '4' or '5' or '6' or '7' or '8' or '9' or '10' or '11' or
            '12' or '13' or '14' or '15' or '16' or '17' or '18' or '19' or '20' or '21' or
            '22' or '23' or '24' or '25' or '26' or '27' or '28' or '29' or '30' or '31'):
        async with state.proxy() as data:
            data['day'] = callback_query.data
        await bot.edit_message_text(chat_id=callback_query.message.chat.id,
                                    message_id=callback_query.message.message_id,
                                    text='Дата твоего рождения', reply_markup=get_birthday_kb())
        await ProfileStatesGroup.input_birthday.set()


@dp.callback_query_handler(state=ProfileStatesGroup.input_month)
async def month_keyboard(callback_query: types.CallbackQuery, state: FSMContext):
    if (callback_query.data == '0 1' or '0 2' or '0 3' or '0 4' or '0 5' or '0 6' or '0 7' or '0 8' or
            '0 9' or '1 0' or '1 1' or '1 2'):
        async with state.proxy() as data:
            data['month'] = callback_query.data

        await bot.edit_message_text(chat_id=callback_query.message.chat.id,
                                    message_id=callback_query.message.message_id,
                                    text='Дата твоего рождения', reply_markup=get_birthday_kb())
        await ProfileStatesGroup.input_birthday.set()


@dp.callback_query_handler(state=ProfileStatesGroup.input_year)
async def year_keyboard(callback_query: types.CallbackQuery, state: FSMContext):
    if (callback_query.data == '1970' or '1972' or '1973' or '1974' or '1975' or '1976' or '1977' or '1978' or
            '1979' or '1980' or '1981' or '1982' or '1983' or '1984' or '1985' or '1986' or '1987' or '1988' or
            '1989' or '1990' or '1991'):
        async with state.proxy() as data:
            data['year'] = callback_query.data
        await bot.edit_message_text(chat_id=callback_query.message.chat.id,
                                    message_id=callback_query.message.message_id,
                                    text='Дата твоего рождения', reply_markup=get_birthday_kb())
        await ProfileStatesGroup.input_birthday.set()


@dp.callback_query_handler(state=ProfileStatesGroup.input_Tashkent_or_other_town)
async def town_keyboard(callback_query: types.CallbackQuery, state: FSMContext):
    if callback_query.data == 'Ташкент':
        await callback_query.message.delete()
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=where_are_you_from)
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=callback_query.data)

        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=district,
                               reply_markup=get_district_kb())
        await ProfileStatesGroup.input_district.set()

    if callback_query.data == 'Другой':
        await callback_query.message.delete()

        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=drugoi)
        await ProfileStatesGroup.input_other_town_and_district.set()

    if callback_query.data == 'back_to_birth':
        await callback_query.message.delete()
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=date_of_birthday,
                               reply_markup=get_birthday_kb())
        await ProfileStatesGroup.input_birthday.set()


@dp.callback_query_handler(state=ProfileStatesGroup.input_district)
async def district_keyboard(callback_query: types.CallbackQuery, state: FSMContext):
    if (
            callback_query.data == 'Алмазар' or callback_query.data == 'Бектемир' or callback_query.data == 'Мирабад' or callback_query.data == 'Мирзо-Улугбек' or callback_query.data == 'Сергели' or
            callback_query.data == 'Чиланзар' or callback_query.data == 'Шайхантаур' or callback_query.data == 'Юнусабад' or callback_query.data == 'Яккасарай' or callback_query.data == 'Яшнабад' or callback_query.data == 'Учтепа'):
        async with state.proxy() as data:
            data['town_and_district'] = f"Ташкент/{callback_query.data}"
        await callback_query.message.delete()
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=district)
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=callback_query.data)
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=education,
                               reply_markup=get_edu_kb())
        await ProfileStatesGroup.input_edu.set()

    if callback_query.data == 'back_to_town':
        await callback_query.message.delete()
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=where_are_you_from,
                               reply_markup=get_town_kb())
        await ProfileStatesGroup.input_Tashkent_or_other_town.set()


@dp.callback_query_handler(state=ProfileStatesGroup.input_edu)
async def edu_keyboard(callback_query: types.CallbackQuery, state: FSMContext):
    if (
            callback_query.data == 'Высшее' or callback_query.data == 'Неполное высшее' or callback_query.data == 'Среднее' or
            callback_query.data == 'Неполное среднее' or callback_query.data == 'Среднее специальное'):
        async with state.proxy() as data:
            data['edu'] = callback_query.data
        await callback_query.message.delete()
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=education)
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=callback_query.data)
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=rus_lang,
                               reply_markup=get_rus_kb())
        await ProfileStatesGroup.input_rus.set()
    if callback_query.data == 'to_town':
        await callback_query.message.delete()
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=where_are_you_from,
                               reply_markup=get_town_kb())
        await ProfileStatesGroup.input_Tashkent_or_other_town.set()


@dp.callback_query_handler(state=ProfileStatesGroup.input_rus)
async def rus_keyboard(callback_query: types.CallbackQuery, state: FSMContext):
    if (
            callback_query.data == 'Отлично' or callback_query.data == 'Хорошо' or callback_query.data == 'Удовлетворительно' or
            callback_query.data == 'Не владею русским языком'):
        async with state.proxy() as data:
            data['rus'] = callback_query.data
        await callback_query.message.delete()
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=rus_lang)
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=callback_query.data)
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=uzb_lang,
                               reply_markup=get_uzb_kb())
        await ProfileStatesGroup.input_uzb.set()
    if callback_query.data == 'back_to_edu':
        await callback_query.message.delete()
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=education,
                               reply_markup=get_edu_kb())
        await ProfileStatesGroup.input_edu.set()


@dp.callback_query_handler(state=ProfileStatesGroup.input_uzb)
async def uzb_keyboard(callback_query: types.CallbackQuery, state: FSMContext):
    if (
            callback_query.data == 'Отлично знаю' or callback_query.data == 'Хорошо знаю' or callback_query.data == 'Удовлетворительно знаю' or
            callback_query.data == 'Не владею узбекским языком'):
        async with state.proxy() as data:
            data['uzb'] = callback_query.data
        await callback_query.message.delete()
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=uzb_lang)
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=callback_query.data)
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=eng_lang,
                               reply_markup=get_eng_kb())
        await ProfileStatesGroup.input_eng.set()
    if callback_query.data == 'back_to_ru':
        await callback_query.message.delete()
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=rus_lang,
                               reply_markup=get_rus_kb())
        await ProfileStatesGroup.input_rus.set()


@dp.callback_query_handler(state=ProfileStatesGroup.input_eng)
async def uzb_keyboard(callback_query: types.CallbackQuery, state: FSMContext):
    if (
            callback_query.data == 'Отлично владею' or callback_query.data == 'Хорошо владею' or callback_query.data == 'Удовлетворительно владею' or
            callback_query.data == 'Не владею английским языком'):
        async with state.proxy() as data:
            data['eng'] = callback_query.data
        await callback_query.message.delete()
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=eng_lang)
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=callback_query.data)
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=experience_msg,
                               reply_markup=get_exp_kb())
        await ProfileStatesGroup.input_experience.set()
    if callback_query.data == 'back_to_uz':
        await callback_query.message.delete()
        await bot.send_message(chat_id=callback_query.message.chat.id,
                               text=uzb_lang,
                               reply_markup=get_uzb_kb())
        await ProfileStatesGroup.input_uzb.set()

    @dp.callback_query_handler(state=ProfileStatesGroup.input_experience)
    async def exp_keyboard(callback_query: types.CallbackQuery, state: FSMContext):
        if callback_query.data == 'Есть':
            await callback_query.message.delete()
            await bot.send_message(chat_id=callback_query.message.chat.id,
                                   text=experience_about)
            await ProfileStatesGroup.experience_describe.set()
        if callback_query.data == 'Нет':
            async with state.proxy() as data:
                data['exp'] = callback_query.data
            await callback_query.message.delete()
            await bot.send_message(chat_id=callback_query.message.chat.id,
                                   text=time_for_call)
            await ProfileStatesGroup.input_day_and_time.set()
        if callback_query.data == 'back_to_eng':
            await callback_query.message.delete()
            await bot.send_message(chat_id=callback_query.message.chat.id,
                                   text=eng_lang,
                                   reply_markup=get_eng_kb())
            await ProfileStatesGroup.input_eng.set()



if __name__ == '__main__':
    scheduler.start()
    executor.start_polling(dp, skip_updates=True)
